import pandas as pd
import openpyxl
import os
from datetime import datetime
import sys

# 创建报告文件
report_file = f'数据质量检查报告_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'

# 同时输出到控制台和文件
class Logger:
    def __init__(self, filename):
        self.terminal = sys.stdout
        self.log = open(filename, 'w', encoding='utf-8')
    
    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)
    
    def flush(self):
        self.terminal.flush()
        self.log.flush()
    
    def close(self):
        self.log.close()

logger = Logger(report_file)
sys.stdout = logger

print("="*80)
print("淘宝商品评论数据质量检查报告".center(80))
print("="*80)
print(f"检查时间: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}")
print(f"报告文件: {report_file}")
print("="*80)

# 查找最新的Excel文件
files = [f for f in os.listdir('.') if f.endswith('_FromTB.xlsx') and not f.startswith('~$')]
if not files:
    print("\n❌ 错误：未找到爬取的数据文件！")
    logger.close()
    sys.stdout = sys.__stdout__
    exit()

# 选择最新的文件
latest_file = max(files, key=lambda f: os.path.getmtime(f))
print(f"\n📁 数据文件信息:")
print(f"   文件名: {latest_file}")
print(f"   文件大小: {os.path.getsize(latest_file)/1024:.2f} KB")
print(f"   修改时间: {datetime.fromtimestamp(os.path.getmtime(latest_file)).strftime('%Y-%m-%d %H:%M:%S')}")

# 用于收集所有异常
all_issues = []

# 读取Excel文件
try:
    wb = openpyxl.load_workbook(latest_file)
    print(f"\n📊 包含的Sheet: {', '.join(wb.sheetnames)}")
    
    for sheet_name in wb.sheetnames:
        print(f"\n{'='*80}")
        print(f"📄 Sheet: {sheet_name}".center(80))
        print(f"{'='*80}")
        
        # 使用pandas读取
        df = pd.read_excel(latest_file, sheet_name=sheet_name)
        
        print(f"\n【一、基本信息】")
        print(f"   ├─ 总行数: {len(df):,} 行")
        print(f"   ├─ 总列数: {len(df.columns)} 列")
        print(f"   └─ 列名: {', '.join(df.columns)}")
        
        print(f"\n【二、数据预览（前3行）】")
        print("-" * 80)
        for idx, row in df.head(3).iterrows():
            print(f"第{idx+1}行:")
            for col in df.columns:
                value = str(row[col])[:50] if pd.notna(row[col]) else 'NaN'
                print(f"   {col}: {value}")
            print("-" * 80)
        
        print(f"\n【三、数据类型】")
        for col, dtype in df.dtypes.items():
            print(f"   {col:20s} : {dtype}")
        
        print(f"\n【四、数据完整性检查】")
        missing = df.isnull().sum()
        missing_percent = (df.isnull().sum() / len(df) * 100).round(2)
        
        if missing.sum() > 0:
            print("   ⚠️ 发现缺失值:")
            for col in df.columns:
                if missing[col] > 0:
                    print(f"      ├─ {col}: {missing[col]} 个 ({missing_percent[col]:.1f}%)")
                    all_issues.append({
                        'sheet': sheet_name,
                        'type': '缺失值',
                        'column': col,
                        'count': missing[col],
                        'percent': missing_percent[col]
                    })
        else:
            print("   ✅ 无缺失值")
        
        print(f"\n【五、重复数据检查】")
        duplicates = df.duplicated().sum()
        if duplicates > 0:
            print(f"   ⚠️ 完全重复行数: {duplicates} ({duplicates/len(df)*100:.2f}%)")
            all_issues.append({
                'sheet': sheet_name,
                'type': '重复数据',
                'count': duplicates,
                'percent': duplicates/len(df)*100
            })
        else:
            print("   ✅ 无重复数据")
        
        print(f"\n【六、空白值检查】")
        has_blank = False
        for col in df.columns:
            if df[col].dtype == 'object':
                empty_count = (df[col].astype(str).str.strip() == '').sum()
                if empty_count > 0:
                    if not has_blank:
                        print("   ⚠️ 发现空白值:")
                        has_blank = True
                    print(f"      ├─ {col}: {empty_count} 个 ({empty_count/len(df)*100:.2f}%)")
        if not has_blank:
            print("   ✅ 无空白值")
        
        if sheet_name == '商品评论':
            print(f"\n{'='*80}")
            print(f"【七、评论内容质量分析】")
            print(f"{'='*80}")
            
            # 检查评论长度
            if '评论内容' in df.columns:
                df['评论长度'] = df['评论内容'].astype(str).str.len()
                
                print(f"\n▶ 评论长度统计:")
                print(f"   ├─ 平均长度: {df['评论长度'].mean():.0f} 字")
                print(f"   ├─ 最短评论: {df['评论长度'].min()} 字")
                print(f"   ├─ 最长评论: {df['评论长度'].max()} 字")
                print(f"   └─ 中位数: {df['评论长度'].median():.0f} 字")
                
                # 检查异常短评论
                short_reviews = df[df['评论长度'] < 5]
                if len(short_reviews) > 0:
                    print(f"\n   ⚠️ 异常短评论(少于5字): {len(short_reviews)} 条 ({len(short_reviews)/len(df)*100:.2f}%)")
                    print(f"\n   【异常短评论详情】")
                    print("   " + "-" * 76)
                    for idx, row in short_reviews.head(10).iterrows():
                        print(f"   序号{idx+1}:")
                        print(f"      用户: {row['用户名']}")
                        print(f"      购买: {row.get('购买记录', 'N/A')}")
                        print(f"      内容: {row['评论内容']}")
                        print(f"      长度: {row['评论长度']} 字")
                        print("   " + "-" * 76)
                    if len(short_reviews) > 10:
                        print(f"   （仅显示前10条，共{len(short_reviews)}条）")
                else:
                    print(f"\n   ✅ 无异常短评论")
                
                # 检查包含商家回复的评论
                seller_reply_reviews = df[df['评论内容'].astype(str).str.contains('商家回复|店家回复|卖家回复', na=False)]
                if len(seller_reply_reviews) > 0:
                    print(f"\n▶ 包含商家回复: {len(seller_reply_reviews)} 条 ({len(seller_reply_reviews)/len(df)*100:.1f}%)")
                    print(f"\n   【包含商家回复的评论示例】")
                    print("   " + "-" * 76)
                    for idx, row in seller_reply_reviews.head(3).iterrows():
                        print(f"   序号{idx+1}:")
                        print(f"      用户: {row['用户名']}")
                        content = row['评论内容'][:200] + '...' if len(row['评论内容']) > 200 else row['评论内容']
                        print(f"      内容: {content}")
                        print("   " + "-" * 76)
                    all_issues.append({
                        'sheet': sheet_name,
                        'type': '包含商家回复',
                        'count': len(seller_reply_reviews),
                        'percent': len(seller_reply_reviews)/len(df)*100
                    })
                else:
                    print(f"\n   ✅ 无商家回复混入")
                
                # 检查包含"\n更多"标记
                more_tag_reviews = df[df['评论内容'].astype(str).str.contains('更多', na=False)]
                if len(more_tag_reviews) > 0:
                    print(f"\n▶ 包含'更多'标记: {len(more_tag_reviews)} 条 ({len(more_tag_reviews)/len(df)*100:.1f}%)")
                    all_issues.append({
                        'sheet': sheet_name,
                        'type': "包含'更多'标记",
                        'count': len(more_tag_reviews),
                        'percent': len(more_tag_reviews)/len(df)*100
                    })
                
                # 检查换行符
                newline_reviews = df[df['评论内容'].astype(str).str.contains('\n|\r', na=False, regex=True)]
                if len(newline_reviews) > 0:
                    print(f"\n▶ 包含换行符: {len(newline_reviews)} 条 ({len(newline_reviews)/len(df)*100:.1f}%)")
                    all_issues.append({
                        'sheet': sheet_name,
                        'type': '包含换行符',
                        'count': len(newline_reviews),
                        'percent': len(newline_reviews)/len(df)*100
                    })
            
            # 检查购买记录完整性
            if '购买记录' in df.columns:
                print(f"\n▶ 购买记录完整性:")
                has_purchase = df['购买记录'].notna() & (df['购买记录'].astype(str).str.strip() != '')
                print(f"   ├─ 包含购买记录: {has_purchase.sum()} 条 ({has_purchase.sum()/len(df)*100:.1f}%)")
                
                missing_purchase = len(df) - has_purchase.sum()
                if missing_purchase > 0:
                    print(f"   └─ ⚠️ 缺失购买记录: {missing_purchase} 条 ({missing_purchase/len(df)*100:.1f}%)")
                    
                    # 显示缺失购买记录的评论
                    no_purchase_reviews = df[~has_purchase]
                    print(f"\n   【缺失购买记录的评论示例】")
                    print("   " + "-" * 76)
                    for idx, row in no_purchase_reviews.head(5).iterrows():
                        print(f"   序号{idx+1}:")
                        print(f"      用户: {row['用户名']}")
                        print(f"      购买记录: {row.get('购买记录', 'N/A')}")
                        content = row['评论内容'][:100] + '...' if len(row['评论内容']) > 100 else row['评论内容']
                        print(f"      内容: {content}")
                        print("   " + "-" * 76)
                    
                    all_issues.append({
                        'sheet': sheet_name,
                        'type': '缺失购买记录',
                        'count': missing_purchase,
                        'percent': missing_purchase/len(df)*100
                    })
                else:
                    print(f"   └─ ✅ 购买记录100%完整")

except Exception as e:
    print(f"\n❌ 读取文件出错: {e}")
    import traceback
    traceback.print_exc()

# 生成总结报告
print(f"\n{'='*80}")
print(f"【检查总结】"。center(80))
print(f"{'='*80}")

if all_issues:
    print(f"\n⚠️ 发现 {len(all_issues)} 类数据质量问题:\n")
    for i, issue in enumerate(all_issues, 1):
        print(f"{i}. [{issue['sheet']}] {issue['type']}")
        if 'column' in issue:
            print(f"   ├─ 列名: {issue['column']}")
        print(f"   ├─ 数量: {issue['count']}")
        print(f"   └─ 占比: {issue.get('percent', 0):.2f}%")
        print()
    
    print(f"\n💡 建议:")
    print(f"   1. 运行数据清洗脚本移除商家回复和无用标记")
    print(f"   2. 清理换行符和多余空白字符")
    print(f"   3. 移除或标记异常短评论")
    print(f"   4. 检查并补充缺失的购买记录")
else:
    print(f"\n✅ 数据质量良好，未发现明显问题！")

print(f"\n{'='*80}")
print(f"检查完成！报告已保存到: {report_file}")
print(f"{'='*80}")

# 关闭日志
logger.close()
sys.stdout = sys.__stdout__

print(f"\n✅ 数据质量检查报告已生成: {report_file}")
